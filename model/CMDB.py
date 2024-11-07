"""Classes and functions for reading/processing/creating a POAM report.

Core functionality:
1) Read in a Excel file that has POAMS stored in the official FedRAMP template
2) TODO:

"""
import datetime
import os
import logging
import coloredlogs
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl
import re
import json
import pandas

SERVER_IP_RE = r"(\S+)\s*\(?.*\)?\s*Ports:\s*((.*)+)"
# SERVER_IP_RE = r"(\S+)\s*Ports:\s*((.*)+)"
STARTS_WITH_AFFECTS_RE = r"\s*Affect.*"

coloredlogs.install(level=logging.INFO,
                    fmt="%(asctime)s %(hostname)s %(name)s %(filename)s line-%(lineno)d %(levelname)s - %(message)s",
                    datefmt='%H:%M:%S')

SPREADSHEET_FIRST_DATA_ROW = 2
# HEADER_ROW_FLAG = True

# This effectively defines the root of the project and so adding ..\, etc is not needed
# in config files,etc
# PROJECT_ROOT_DIR = os.path.dirname(os.path.dirname(__file__))

# Add script directory to the path to allow searching for modules
# sys.path.insert(0, PROJECT_ROOT_DIR)

YEAR_MONTH_DAY = datetime.datetime.now().strftime("%Y%m%d")
ENVIRONMENT = "L5"

FEDRAMP_ROOT_DIR = "C:\\Users\\dhartman\\Documents\\FedRAMP\\CMDB\\"
CMDB_SPREADSHEET_FILENAME = "{environment}\\CMDB-{environment}-Master.xlsx".format(
        environment=ENVIRONMENT)
CMDB_RESULTS_SPREADSHEET_FULL_PATH = os.path.join(FEDRAMP_ROOT_DIR, CMDB_SPREADSHEET_FILENAME)
CMDB_WORKSHEET = "CMDB-INVENTORY"
SCAN_TARGET_POLICY_SHEET = "all_scans_policies_targets"

SYSTEM_INVENTORY_WORKSHEET = "SystemInventory"
SERVER_ALIAS_MAPPING_WORKSHEET = "NAME-ALIAS-MAPPING"
SERVER_ALIAS_MAPPING_START_ROW = 2

CMDB_EXCEL_TEMPLATE_FILENAME = "PTC-CMDB-Inventory-Working-Template.xlsx"

FEDRAMP_SYSTEM_INVENTORY_SHEET = "Inventory"
FEDRAMP_SYSTEM_INVENTORY_DATA_ROW_START = 3

CMDB_INVENTORY_SHEET = "CMDB"
CMDB_INVENTORY_SHEET_DATA_ROW_START = 2

FEDRAMP_SYSTEM_INVENTORY_DATA_ROW_START

CMDB_TEMPLATE_EXCEL_PATH = os.path.join(FEDRAMP_ROOT_DIR, CMDB_EXCEL_TEMPLATE_FILENAME)

# Output Files
CMDB_OUTPUT_FILE_JSON = "{environment}\\PTC-{environment}-cmdb.json".format(
        environment=ENVIRONMENT,
        year_month_day=YEAR_MONTH_DAY)
CMDB_RESULTS_JSON_FULL_PATH = os.path.join(FEDRAMP_ROOT_DIR, CMDB_OUTPUT_FILE_JSON)

CMDB_FIELDS = ['ID',
               'PRIMARY_IP_ADDRESS',
               'NAME',
               'ADDITIONAL_IP_ADDRESSES',
               'ENVIRONMENT',
               'FUNCTION',
               'SYSTEM_ADMINISTRATOR_OWNER',
               'APPLICATION_ADMINISTRATOR_OWNER',
               'UNIQUE_ASSET_IDENTIFIER',
               'IPV4_OR_IPV6_ADDRESS',
               'VIRTUAL',
               'PUBLIC',
               'DNS_NAME_OR_URL',
               'NETBIOS_NAME',
               'MAC_ADDRESS',
               'AUTHENTICATED_SCAN',
               'BASELINE_CONFIGURATION_NAME',
               'OS_NAME_AND_VERSION',
               'LOCATION',
               'ASSET_TYPE',
               'HARDWARE_MAKE_MODEL',
               'IN_LATEST_SCAN',
               'SOFTWARE_DATABASE_VENDOR',
               'SOFTWARE_DATABASE_NAME_VERSION',
               'PATCH_LEVEL',
               'COMMENTS',
               'SERIAL_NUMBER_ASSET_TAG_NUMBER',
               'VLAN_NETWORK_ID'
               ]

SCAN_POLICY_TARGET = ["SCAN_NAME",
                      "POLICY_NAME",
                      "CREDENTIALS",
                      "TARGETS"]


def datetime_default(obj):
    if isinstance(obj, (datetime.date, datetime.time)):
        return obj.isoformat()


def clean_cell(cell):
    if type(cell) == str:
        return cell.strip()
    elif cell is None:
        return ""

    return cell


class CMDBInventoryItem:
    def __init__(self,
                 id,
                 primary_ip_address,
                 name,
                 additional_ip_addresses,
                 environment,
                 function,
                 system_administrator_owner,
                 application_administrator_owner,
                 unique_asset_identifier,
                 ipv4_or_ipv6_address,
                 virtual,
                 public,
                 dns_name_or_url,
                 netbios_name,
                 mac_address,
                 authenticated_scan,
                 baseline_configuration_name,
                 os_name_and_version,
                 location,
                 asset_type,
                 hardware_make_model,
                 in_latest_scan,
                 software_database_vendor,
                 software_database_name_version,
                 patch_level,
                 comments,
                 serial_number_asset_tag_number,
                 vlan_network_id
                 ):
        self.id = id
        self.primary_ip_address = primary_ip_address
        self.name = name
        self.additional_ip_addresses = additional_ip_addresses
        self.environment = environment
        self.function = function
        self.system_administrator_owner = system_administrator_owner
        self.application_administrator_owner = application_administrator_owner
        self.unique_asset_identifier = unique_asset_identifier
        self.ipv4_or_ipv6_address = ipv4_or_ipv6_address
        self.virtual = virtual
        self.public = public
        self.dns_name_or_url = dns_name_or_url
        self.netbios_name = netbios_name
        self.mac_address = mac_address
        self.authenticated_scan = authenticated_scan
        self.baseline_configuration_name = baseline_configuration_name
        self.os_name_and_version = os_name_and_version
        self.location = location
        self.asset_type = asset_type
        self.hardware_make_model = hardware_make_model
        self.in_latest_scan = in_latest_scan
        self.software_database_vendor = software_database_vendor
        self.software_database_name_version = software_database_name_version
        self.patch_level = patch_level
        self.comments = comments
        self.serial_number_asset_tag_number = serial_number_asset_tag_number
        self.vlan_network_id = vlan_network_id
        self.scans = []
        self.ip_array = [primary_ip_address]
        for additional_ip in additional_ip_addresses.split():
            self.ip_array.append(additional_ip)
        # create an array of name, type and value alias mappings
        self.name_alias_mapping = []
        primary_ip_dict_to_add = dict(NAME=self.name, TYPE="PRIMARY_IP", VALUE=self.primary_ip_address)
        self.name_alias_mapping.append(primary_ip_dict_to_add)
        for ips in self.ip_array:
            ips_dict_to_add = dict(NAME=self.name, TYPE="IP", VALUE=ips)
            self.name_alias_mapping.append(ips_dict_to_add)
        if self.dns_name_or_url != "":
            dns_dict_to_add = dict(NAME=self.name, TYPE="DNS", VALUE=self.dns_name_or_url.upper())
            self.name_alias_mapping.append(dns_dict_to_add)
        if self.netbios_name != "":
            netbios_dict_to_add = dict(NAME=self.name, TYPE="NETBIOS", VALUE=self.netbios_name.upper())
            self.name_alias_mapping.append(netbios_dict_to_add)
        if self.mac_address != "":
            mac_address_dict_to_add = dict(NAME=self.name, TYPE="MAC", VALUE=self.mac_address.upper())
            self.name_alias_mapping.append(mac_address_dict_to_add)

    def __str__(self):
        return_str = ""
        try:
            return_str = "CMDB ID {cmdb_id} Name-IP: {name} - {ip}".format(cmdb_id=self.unique_asset_identifier,
                                                                           name=self.name,
                                                                           ip=self.primary_ip_address)
        except Exception as e:
            logging.error("Could not print POAM")
        return return_str

    def get_cmdb_item_dict(self):
        cmdb_item_dict = dict(ID=self.id,
                              NAME=self.name,
                              ENVIRONMENT=self.environment,
                              IPS=self.ip_array,
                              NAME_ALIAS_MAPPING=self.name_alias_mapping
                              )
        return cmdb_item_dict


class CMDB:
    """
    Contains all CMDB System inventory objects.  When CMDB object is created, if no spreadsheet name is provided, use default constant and load into POAM objects.

    - process target scan details to identify which tenable scans scan the cmdb item

    Attributes
    ----------
    report_year_month: string
        year and month for reporting of inventory

    Methods
    -------
    load_cmdb_excel(environment, in_cmdb_workbook, cmdb_worksheet_name)
        Creates an instance of CMDB based on an excel file
    """
    tab_separator = "\t"
    comma_separator = ","

    def __init__(self,
                 environment,
                 cmdb_year_month_period,
                 cmdb_items=None):
        """ If called with no file name, load default spreadsheet for the environment
        """
        self.environment = environment
        self.cmdb_year_month_period = cmdb_year_month_period
        self.cmdb_report_create_date = datetime.datetime.now()
        if cmdb_items is None:
            self.cmdb_items = []
            self.results_count = 0
            # self.name_alias_mapping = []
        else:
            self.results_count = len(cmdb_items)
            self.cmdb_items = cmdb_items

    @classmethod
    def load_cmdb_excel(cls, environment, in_cmdb_workbook, cmdb_worksheet_name):
        if in_cmdb_workbook.endswith(".xlsx") or \
                in_cmdb_workbook.endswith(".xlsm") or \
                in_cmdb_workbook.endswith(".xls"):
            logging.info("processing excel file [%s]", in_cmdb_workbook)
        else:
            logging.error("[%s] is not an excel file file", in_cmdb_workbook)
            raise NameError(in_cmdb_workbook)
        logging.debug("cmdb excel processing excel file [%s]", in_cmdb_workbook)
        cmdb_wb = openpyxl.load_workbook(in_cmdb_workbook)
        cmdb_ws = cmdb_wb[cmdb_worksheet_name]
        cmdb_results_out = []
        line_number = 0
        for raw_row in cmdb_ws.iter_rows(min_row=SPREADSHEET_FIRST_DATA_ROW, max_col=100, values_only=True):
            row = [clean_cell(field) for field in raw_row]
            if row[CMDB_FIELDS.index('UNIQUE_ASSET_IDENTIFIER')] is None:
                continue
            line_number += 1
            unique_asset_identifier = row[CMDB_FIELDS.index('UNIQUE_ASSET_IDENTIFIER')].upper()
            cmdb_item = CMDBInventoryItem(id=row[CMDB_FIELDS.index('ID')],
                                          primary_ip_address=row[CMDB_FIELDS.index('PRIMARY_IP_ADDRESS')],
                                          name=row[CMDB_FIELDS.index('NAME')],
                                          additional_ip_addresses=row[CMDB_FIELDS.index('ADDITIONAL_IP_ADDRESSES')],
                                          environment=row[CMDB_FIELDS.index('ENVIRONMENT')],
                                          function=row[CMDB_FIELDS.index('FUNCTION')],
                                          system_administrator_owner=row[
                                              CMDB_FIELDS.index('SYSTEM_ADMINISTRATOR_OWNER')],
                                          application_administrator_owner=row[
                                              CMDB_FIELDS.index('APPLICATION_ADMINISTRATOR_OWNER')],
                                          unique_asset_identifier=row[CMDB_FIELDS.index('UNIQUE_ASSET_IDENTIFIER')],
                                          ipv4_or_ipv6_address=row[CMDB_FIELDS.index('IPV4_OR_IPV6_ADDRESS')],
                                          virtual=row[CMDB_FIELDS.index('VIRTUAL')],
                                          public=row[CMDB_FIELDS.index('PUBLIC')],
                                          dns_name_or_url=row[CMDB_FIELDS.index('DNS_NAME_OR_URL')],
                                          netbios_name=row[CMDB_FIELDS.index('NETBIOS_NAME')],
                                          mac_address=row[CMDB_FIELDS.index('MAC_ADDRESS')],
                                          authenticated_scan=row[CMDB_FIELDS.index('AUTHENTICATED_SCAN')],
                                          baseline_configuration_name=row[
                                              CMDB_FIELDS.index('BASELINE_CONFIGURATION_NAME')],
                                          os_name_and_version=row[CMDB_FIELDS.index('OS_NAME_AND_VERSION')],
                                          location=row[CMDB_FIELDS.index('LOCATION')],
                                          asset_type=row[CMDB_FIELDS.index('ASSET_TYPE')],
                                          hardware_make_model=row[CMDB_FIELDS.index('HARDWARE_MAKE_MODEL')],
                                          in_latest_scan=row[CMDB_FIELDS.index('IN_LATEST_SCAN')],
                                          software_database_vendor=row[CMDB_FIELDS.index('SOFTWARE_DATABASE_VENDOR')],
                                          software_database_name_version=row[
                                              CMDB_FIELDS.index('SOFTWARE_DATABASE_NAME_VERSION')],
                                          patch_level=row[CMDB_FIELDS.index('PATCH_LEVEL')],
                                          comments=row[CMDB_FIELDS.index('COMMENTS')],
                                          serial_number_asset_tag_number=row[
                                              CMDB_FIELDS.index('SERIAL_NUMBER_ASSET_TAG_NUMBER')],
                                          vlan_network_id=row[CMDB_FIELDS.index('VLAN_NETWORK_ID')]
                                          )
            cmdb_results_out.append(cmdb_item)
        return cls(environment=environment, cmdb_year_month_period="2020-June", cmdb_items=cmdb_results_out)

    def get_cmdb_item(self, in_value):
        """Pass in a IP, DNS Name, name, etc and return the CMDBItems
        """
        return_cmdb_item = None
        for cmdb_item in self.cmdb_items:
            for name_alias_map in cmdb_item.name_alias_mapping:
                if in_value.upper() == name_alias_map['VALUE'].upper():
                    # return the full CMDBItem
                    return_cmdb_item = cmdb_item
                    # return_name_value = name_alias_map[0]
        return return_cmdb_item

    def copy_name_alias_mapping_to_clipboard(self):
        mapping = []
        for item in self.cmdb_items:
            for mapping_item in item.name_alias_mapping:
                mapping.append([mapping_item[0].upper(), mapping_item[1][1]])
        df = pandas.DataFrame(mapping)
        df.to_clipboard(index=False, header=False)

    def update_scan_target_info(self, scan_target_workbook, scan_target_worksheet_name):
        logging.info("Getting Nessus Scan/Target details")
        scan_policy_target_wb = openpyxl.load_workbook(scan_target_workbook)
        scan_policy_target_ws = scan_policy_target_wb[scan_target_worksheet_name]
        line_number = 0
        scan_target = []
        for row in scan_policy_target_ws.iter_rows(min_row=2, max_col=100, values_only=True):
            if row[SCAN_POLICY_TARGET.index('SCAN_NAME')] is None:
                continue
            line_number += 1

            stripped_scan_result_row = [field.strip() if type(field) == str else str(field) for field in row]

            scan_name = stripped_scan_result_row[SCAN_POLICY_TARGET.index('SCAN_NAME')]
            policy_name = stripped_scan_result_row[SCAN_POLICY_TARGET.index('POLICY_NAME')]
            credentials = stripped_scan_result_row[SCAN_POLICY_TARGET.index('CREDENTIALS')]
            target_list = stripped_scan_result_row[SCAN_POLICY_TARGET.index('TARGETS')]
            targets = [item.strip() for item in target_list.split(',')]
            for target in targets:
                # logging.info("updating target [%s]", target)
                cmdb_item = self.get_cmdb_item(target)
                if cmdb_item is not None:
                    cmdb_item.scans.append([scan_name, policy_name, credentials, target])
                else:
                    logging.warning("For scan [%s] and target [%s] there was no CMDB Item found", scan_name, target)

                # scan_target.append([scan_name, policy_name, credentials, target])
            # print(str(scan_target))
            # for each IP/Name, update the CMDB Item with the scan information
            # df = pandas.DataFrame(scan_target)
            # df.to_clipboard(index=False, header=False)

    def create_cmdb_json_file(self, out_json_file):
        logging.info("create cmdb results JSON output to file [%s]", out_json_file)
        current_date_time = datetime.datetime.now().strftime('%B %d, %Y %I:%M:%S EDT')
        cmdb_items_array = []
        for cmdb_item_entry in self.cmdb_items:
            cmdb_items_array.append(cmdb_item_entry.get_cmdb_item_dict())

        out_dictionary = dict(ENVIRONMENT=self.environment,
                              SCAN_RESULT_COUNT=self.results_count,
                              YEAR_MONTH_PERIOD=self.cmdb_year_month_period,
                              REPORT_GEN_DATETIME=current_date_time,  # self.in_scan_date, # datetime.now(),
                              CMDB_ITEMS=cmdb_items_array)
        with open(out_json_file, 'w') as json_out_handle:
            json.dump(out_dictionary, json_out_handle, default=datetime_default)
        return out_dictionary

    def get_name_alias_mapping_array(self):
        """Returns a 2 dimensional array that includes the item name, alias type and alias value """
        name_alias_mapping_return = []
        for cmdb_item in self.cmdb_items:
            name = cmdb_item.name
            for alias_type_value in cmdb_item.name_alias_mapping:
                alias_type = alias_type_value['TYPE']
                alias_value = alias_type_value['VALUE']
                name_alias_mapping_return.append([name, alias_type, alias_value])
        return name_alias_mapping_return

    def create_cmdb_excel_table(self, in_worksheet):
        cdmb_table_data = []
        CMDB_TABLE_START_COLUMN = "A"
        CMDB_TABLE_END_COLUMN = "AB"
        CMDB_TABLE_START_ROW = "1"
        # tab = openpyxl.worksheet.table.Table(displayName="CMDB")
        # Add a default style with striped rows and banded columns
        cmdb_inventory_table = Table(displayName="CMDB_TABLE")
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        # in_table.tableStyleInfo = style

    def create_cmdb_inventory_spreadsheet(self, out_spreadsheet_name):
        CMDB_TABLE_NAME = "CMDB"
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        logging.info("Creating excel CMDB Inventory spreadsheet [%s] from template [%s]", out_spreadsheet_name, CMDB_TEMPLATE_EXCEL_PATH)
        cmdb_wb = openpyxl.load_workbook(filename=CMDB_TEMPLATE_EXCEL_PATH)
        # create the main CMDB inventory spreadsheet
        cmdb_inventory_worksheet = cmdb_wb[CMDB_INVENTORY_SHEET]
        cmdb_inventory_worksheet.append(CMDB_FIELDS)
        row_count = 1
        for cmdb_system in self.cmdb_items:
            row_count += 1
            row_to_add = [cmdb_system.id,
                          cmdb_system.primary_ip_address,
                          cmdb_system.name,
                          cmdb_system.additional_ip_addresses,
                          cmdb_system.environment,
                          cmdb_system.function,
                          cmdb_system.system_administrator_owner,
                          cmdb_system.application_administrator_owner,
                          cmdb_system.unique_asset_identifier,
                          cmdb_system.ipv4_or_ipv6_address,
                          cmdb_system.virtual,
                          cmdb_system.public,
                          cmdb_system.dns_name_or_url,
                          cmdb_system.netbios_name,
                          cmdb_system.mac_address,
                          cmdb_system.authenticated_scan,
                          cmdb_system.baseline_configuration_name,
                          cmdb_system.os_name_and_version,
                          cmdb_system.location,
                          cmdb_system.asset_type,
                          cmdb_system.hardware_make_model,
                          cmdb_system.in_latest_scan,
                          cmdb_system.software_database_vendor,
                          cmdb_system.software_database_name_version,
                          cmdb_system.patch_level,
                          cmdb_system.comments,
                          cmdb_system.serial_number_asset_tag_number,
                          cmdb_system.vlan_network_id
                          ]
            cmdb_inventory_worksheet.append(row_to_add)
        table_ref = "A1:AB{row_count}".format(row_count=row_count)
        cmdb_excel_table = Table(displayName=CMDB_TABLE_NAME, ref=table_ref)
        cmdb_excel_table.tableStyleInfo = style
        cmdb_inventory_worksheet.add_table(cmdb_excel_table)

        # cmdb_inventory_row = CMDB_INVENTORY_SHEET_DATA_ROW_START - 1
        # for cmdb_inventory_item in self.cmdb_items:
        #     cmdb_inventory_row += 1
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('ID') + 1,
        #                                   value=cmdb_inventory_item.id)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('PRIMARY_IP_ADDRESS') + 1,
        #                                   value=cmdb_inventory_item.primary_ip_address)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('NAME') + 1,
        #                                   value=cmdb_inventory_item.name)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row,
        #                                   column=CMDB_FIELDS.index('ADDITIONAL_IP_ADDRESSES') + 1,
        #                                   value=cmdb_inventory_item.additional_ip_addresses)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('ENVIRONMENT') + 1,
        #                                   value=cmdb_inventory_item.environment)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('FUNCTION') + 1,
        #                                   value=cmdb_inventory_item.function)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row,
        #                                   column=CMDB_FIELDS.index('SYSTEM_ADMINISTRATOR_OWNER') + 1,
        #                                   value=cmdb_inventory_item.system_administrator_owner)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row,
        #                                   column=CMDB_FIELDS.index('APPLICATION_ADMINISTRATOR_OWNER') + 1,
        #                                   value=cmdb_inventory_item.application_administrator_owner)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row,
        #                                   column=CMDB_FIELDS.index('UNIQUE_ASSET_IDENTIFIER') + 1,
        #                                   value=cmdb_inventory_item.unique_asset_identifier)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('IPV4_OR_IPV6_ADDRESS') + 1,
        #                                   value=cmdb_inventory_item.ipv4_or_ipv6_address)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('VIRTUAL') + 1,
        #                                   value=cmdb_inventory_item.virtual)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('PUBLIC') + 1,
        #                                   value=cmdb_inventory_item.public)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('DNS_NAME_OR_URL') + 1,
        #                                   value=cmdb_inventory_item.dns_name_or_url)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('NETBIOS_NAME') + 1,
        #                                   value=cmdb_inventory_item.netbios_name)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('MAC_ADDRESS') + 1,
        #                                   value=cmdb_inventory_item.mac_address)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('AUTHENTICATED_SCAN') + 1,
        #                                   value=cmdb_inventory_item.authenticated_scan)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row,
        #                                   column=CMDB_FIELDS.index('BASELINE_CONFIGURATION_NAME') + 1,
        #                                   value=cmdb_inventory_item.baseline_configuration_name)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('OS_NAME_AND_VERSION') + 1,
        #                                   value=cmdb_inventory_item.os_name_and_version)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('LOCATION') + 1,
        #                                   value=cmdb_inventory_item.location)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('ASSET_TYPE') + 1,
        #                                   value=cmdb_inventory_item.asset_type)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('HARDWARE_MAKE_MODEL') + 1,
        #                                   value=cmdb_inventory_item.hardware_make_model)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('IN_LATEST_SCAN') + 1,
        #                                   value=cmdb_inventory_item.in_latest_scan)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row,
        #                                   column=CMDB_FIELDS.index('SOFTWARE_DATABASE_VENDOR') + 1,
        #                                   value=cmdb_inventory_item.software_database_vendor)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row,
        #                                   column=CMDB_FIELDS.index('SOFTWARE_DATABASE_NAME_VERSION') + 1,
        #                                   value=cmdb_inventory_item.software_database_name_version)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('PATCH_LEVEL') + 1,
        #                                   value=cmdb_inventory_item.patch_level)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('COMMENTS') + 1,
        #                                   value=cmdb_inventory_item.comments)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row,
        #                                   column=CMDB_FIELDS.index('SERIAL_NUMBER_ASSET_TAG_NUMBER') + 1,
        #                                   value=cmdb_inventory_item.serial_number_asset_tag_number)
        #     cmdb_inventory_worksheet.cell(row=cmdb_inventory_row, column=CMDB_FIELDS.index('VLAN_NETWORK_ID') + 1,
        #                                   value=cmdb_inventory_item.vlan_network_id)

        # create tab in System Inventory format
        # cmdb_fedramp_inventory_worksheet = cmdb_wb[FEDRAMP_SYSTEM_INVENTORY_SHEET]
        # cmdb_sysytem_inventory_row = FEDRAMP_SYSTEM_INVENTORY_DATA_ROW_START - 1
        # for cmdb_inventory_item in self.cmdb_items:
        #     cmdb_sysytem_inventory_row += 1
        #     cmdb_fedramp_inventory_worksheet.cell(row=cmdb_sysytem_inventory_row, column=1,
        #                                           value=cmdb_inventory_item.name)
        #
        # # create the name/alias mapping (NAME->IP, NAME->DNS, etc)
        # cmdb_name_alias_mapping_worksheet = cmdb_wb[SERVER_ALIAS_MAPPING_WORKSHEET]
        # cmdb_name_alias_map_row = SERVER_ALIAS_MAPPING_START_ROW - 1
        # for cmdb_inventory_item in self.cmdb_items:
        #     name_alias_maps = cmdb_inventory_item.name_alias_mapping
        #     for name_alias_map in name_alias_maps:
        #         cmdb_name_alias_map_row += 1
        #         cmdb_name_alias_mapping_worksheet.cell(row=cmdb_name_alias_map_row, column=1,
        #                                                value=cmdb_inventory_item.name)
        #         cmdb_name_alias_mapping_worksheet.cell(row=cmdb_name_alias_map_row, column=2,
        #                                                value=name_alias_map['TYPE'])
        #         cmdb_name_alias_mapping_worksheet.cell(row=cmdb_name_alias_map_row, column=3,
        #                                                value=name_alias_map['VALUE'])

        # create scan name, name mapping (states which inventory items are scanned in each job)

        logging.info("Saving CMDB Excel to [%s]", out_spreadsheet_name)
        cmdb_wb.save(out_spreadsheet_name)


if __name__ == "__main__":
    logging.info("Testing POAM_Report class")
    current_cmdb = CMDB.load_cmdb_excel(environment="L5", in_cmdb_workbook=CMDB_RESULTS_SPREADSHEET_FULL_PATH,
                                        cmdb_worksheet_name=CMDB_WORKSHEET)
    for item in current_cmdb.cmdb_items:
        logging.debug("cmdb item [%s] name/alias mapping [%s]", item.name, str(item.name_alias_mapping))
    current_cmdb.update_scan_target_info(scan_target_workbook=CMDB_RESULTS_SPREADSHEET_FULL_PATH,
                                         scan_target_worksheet_name=SCAN_TARGET_POLICY_SHEET)
    # Create new spreadsheet with CMDB, System Inventory (in the FedRAMP format), NAME_ALIAS Mapping, and NAME_SCAN mapping worksheets
    test_ip = "10.188.240.135"
    cmdb_item = current_cmdb.get_cmdb_item(test_ip)

    logging.info("name for alias [%s]: %s", test_ip, str(cmdb_item))
    dict_obj = cmdb_item.get_cmdb_item_dict()

    #    for cmdb_entry in current_cmdb.cmdb_items:

    out_json_file = os.path.join(FEDRAMP_ROOT_DIR, "test_cmdb_out.json")
    # This also returns the dictionary for use in subequent processing
    # cmdb_dictionary = current_cmdb.create_cmdb_json_file(out_json_file=out_json_file)
    # name_alias_map = current_cmdb.get_name_alias_mapping_array()
    # logging.info("map [%s]", str(name_alias_map))
    CMDB_OUTPUT_FILE_PATH = os.path.join(FEDRAMP_ROOT_DIR, "L5//", "PTC-CS-L5-Inventory-07232020.xlsx")
    current_cmdb.create_cmdb_inventory_spreadsheet(out_spreadsheet_name=CMDB_OUTPUT_FILE_PATH)
