"""Classes and functions for reading/processing/creating a POAM report.

POAM Report long description

Attributes
----------
in_year : str
   The full year for the POAM Report (e.g. 2020)
in_month : str
   The month number for the POAM Report (e.g. 5)
in_report_name : str
   The general name for the report
poam_report_create_date : datetime
   Date and time the report was created.  This is included in the metadata in the JSON formatted output
results_count : integer
   The total count of POAM in the report.  (default = 0)
TODO: Need to turn the poams object back into an array.  It is currently a dictionary that has the poam_id as the key, but if there is no poam_id assigned yet, this won't be known.
poams : dictionary
   A dictionary that is keyed on the POAM_ID

Methods
-------
process_poam_excel_file()
   Creates a POAM_Report ojbect based on an excel workbook that has POAMS in the FedRAMP template/format.
process_json_file()
   Creates POAM_Report object from a previously created JSON file of a POAM_Report object
create_json_poam_results_file(out_json_file)
   Creates a JSON export of the POAM report object including the POAMS exported as a JSON dictionary (See POAM Class)

TODO-Items Core functionality:
DONE: Read in a Excel file that has POAMS stored in the official FedRAMP template
TODO: Read in a JSON file of current POAMS
TODO: Process the asset identifier and create data structure that has the CMDB Inventory Item and Ports
TODO: Create single POAM object from vulnerability scan results data
TODO: Create full POAM report object based on complete set of vulnerability scan results
TODO: Create Excel report from POAM_Report object
TODO: Create JSON file from POAM_Report object
TODO: Compare 2 reports (e.g. Current and previously submitted) and report differences
TODO: When processing a current POAM report, perform Lookup on previous POAM report to get POAM IDs, etc
TODO: Add ability to generate multiple notes/comments against a POAM and also generate a report of the notes.
TODO: Create and return dictionary object that can be used to easily create JSON file

DONE: Split out POAM and POAM_Report classes

"""
from datetime import datetime
import os
import logging
import coloredlogs
import openpyxl
import re
import pprint
import json
import datetime
from model.POAM import POAM
from model.ScanResults import ScanResults

SERVER_IP_RE = r"(\S+)\s*\(?.*\)?\s*Ports:\s*((.*)+)"
# SERVER_IP_RE = r"(\S+)\s*Ports:\s*((.*)+)"
STARTS_WITH_AFFECTS_RE = r"\s*Affect.*"

coloredlogs.install(level=logging.INFO,
                    fmt="%(asctime)s %(hostname)s %(name)s %(filename)s line-%(lineno)d %(levelname)s - %(message)s",
                    datefmt='%H:%M:%S')


def datetime_default(obj):
    if isinstance(obj, (datetime.date, datetime.day)):
        return obj.isoformat()


POAM_FIELDS = [
        'POAM_ID',
        'CONTROLS',
        'WEAKNESS_NAME',
        'WEAKNESS_DESCRIPTION',
        'WEAKNESS_DETECTOR_SOURCE',
        'WEAKNESS_SOURCE_IDENTIFIER',
        'ASSET_IDENTIFIER',
        'POINT_OF_CONTACT',
        'RESOURCES_REQUIRED',
        'OVERALL_REMEDIATION_PLAN',
        'ORIGINAL_DETECTION_DATE',
        'SCHEDULED_COMPLETION_DATE',
        'PLANNED_MILESTONES',
        'MILESTONE_CHANGES',
        'STATUS_DATE',
        'VENDOR_DEPENDENCY',
        'LAST_VENDOR_CHECK_IN_DATE',
        'VENDOR_DEPENDENT_PRODUCT_NAME',
        'ORIGINAL_RISK_RATING',
        'ADJUSTED_RISK_RATING',
        'RISK_ADJUSTMENT',
        'FALSE_POSITIVE',
        'OPERATIONAL_REQUIREMENT',
        'DEVIATION_RATIONALE',
        'SUPPORTING_DOCUMENTS',
        'COMMENTS',
        'AUTO_APPROVE',
        'IP_PORT'
]


def clean_cell(cell):
    if type(cell) == str:
        return cell.strip()
    elif cell is None:
        return ""

    return cell


class POAMReport:
    tab_separator = "\t"
    comma_separator = ","
    FRM_ENVIRONMENT = "FRM"
    L5_ENVIRONMENT = "L5"
    YEAR_MONTH_DAY = datetime.datetime.now().strftime("%Y%m%d")
    # 3 character month abbreviation
    MONTH_ABBREV = datetime.datetime.now().strftime("%b")
    # Month number
    MONTH_NUM = datetime.datetime.now().strftime("%m")
    YEAR = datetime.datetime.now().strftime("%Y")

    # C:\Users\dhartman\Documents\FedRAMP\Continuous Monitoring\Vulnerability-Scanning\L5\05 - MAY\PTC_L5_Weekly_Scan_Review_05312020.xlsx
    #    FEDRAMP_DIR = "C:\\Users\\dhartman\\Documents\\FedRAMP\\Continuous Monitoring\\Vulnerability-Scanning\\{environment}\\".format(environment=FRM_ENVIRONMENT)
    FEDRAMP_POAM_TEMPLATE_DIR = "C:\\Users\\dhartman\\Documents\\FedRAMP\\Continuous Monitoring\\POAM\\"

    BLANK_FRM_FEDRAMP_FORM = "FedRAMP-{environment}-POAM-Blank.xlsm".format(environment=FRM_ENVIRONMENT)
    FRM_POAM_TEMPLATE_FORM_FULL_PATH = os.path.join(FEDRAMP_POAM_TEMPLATE_DIR, BLANK_FRM_FEDRAMP_FORM)

    BLANK_L5_FEDRAMP_FORM = "FedRAMP-{environment}-POAM-Blank.xlsm".format(environment=L5_ENVIRONMENT)
    L5_POAM_TEMPLATE_FORM_FULL_PATH = os.path.join(FEDRAMP_POAM_TEMPLATE_DIR, BLANK_L5_FEDRAMP_FORM)

    BLANK_POAM_WORKSHEET = "BLANK_POAM"
    BLANK_POAM_FORM_FIRST_ROW = 6
    POAM_DATE_FIELD = "D3"
    AFFECTED_HOST_COUNT_STRING = "Affects {host_count} Host(s):\n"
    AFFECTED_SERVER_STRING = "{resource} ({ip_address}) Ports:{port}\n"

    def __init__(self,
                 in_year="",
                 in_month="",
                 report_name="",
                 poams=None
                 ):
        if poams is None:
            self.poams = []
            self.results_count = 0
        else:
            self.results_count = len(poams)
            self.poams = poams
        self.report_year = in_year
        self.report_month = in_month
        self.report_name = report_name
        self.poam_report_create_date = datetime.datetime.now()

    @classmethod
    def poam_from_excel_report(cls, in_year, in_month, in_poam_excel_file, in_open_poam_worksheet_name):
        if in_poam_excel_file.endswith(".xlsx") or \
                in_poam_excel_file.endswith(".xlsm") or \
                in_poam_excel_file.endswith(".xls"):
            logging.info("processing excel file [%s]", in_poam_excel_file)
        else:
            logging.error("[%s] is not an excel file file", in_poam_excel_file)
            raise NameError(in_poam_excel_file)
        logging.debug("process_poam_excel_file processing excel file [%s]", in_poam_excel_file)
        poam_wb = openpyxl.load_workbook(in_poam_excel_file)
        open_poam_ws = poam_wb[in_open_poam_worksheet_name]
        poam_entries = []
        line_number = 0
        for row in open_poam_ws.iter_rows(min_row=6, max_col=100, values_only=True):
            # First replace None values with empty strings
            # stripped_poam_row = [field.strip() if type(field) == str else str(field) for field in row]
            stripped_poam_row = [clean_cell(field) for field in row]
            #            row_without_none = [(u"" if cell.value is None else unicode(cell.value)) for cell in rownum]
            if row[POAM_FIELDS.index('POAM_ID')] is None:
                continue
            line_number += 1
            poam_id = str(stripped_poam_row[POAM_FIELDS.index('POAM_ID')]).upper()
            controls = stripped_poam_row[POAM_FIELDS.index('CONTROLS')]
            weakness_name = stripped_poam_row[POAM_FIELDS.index('WEAKNESS_NAME')]
            weakness_description = stripped_poam_row[POAM_FIELDS.index('WEAKNESS_DESCRIPTION')]
            weakness_detector_source = stripped_poam_row[POAM_FIELDS.index('WEAKNESS_DETECTOR_SOURCE')]
            weakness_source_identifier = stripped_poam_row[POAM_FIELDS.index('WEAKNESS_SOURCE_IDENTIFIER')]
            asset_identifier = stripped_poam_row[POAM_FIELDS.index('ASSET_IDENTIFIER')]
            point_of_contact = stripped_poam_row[POAM_FIELDS.index('POINT_OF_CONTACT')]
            resources_required = stripped_poam_row[POAM_FIELDS.index('RESOURCES_REQUIRED')]
            overall_remediation_plan = stripped_poam_row[POAM_FIELDS.index('OVERALL_REMEDIATION_PLAN')]
            original_detection_date = stripped_poam_row[POAM_FIELDS.index('ORIGINAL_DETECTION_DATE')]
            scheduled_completion_date = stripped_poam_row[POAM_FIELDS.index('SCHEDULED_COMPLETION_DATE')]
            planned_milestones = stripped_poam_row[POAM_FIELDS.index('PLANNED_MILESTONES')]
            milestone_changes = stripped_poam_row[POAM_FIELDS.index('MILESTONE_CHANGES')]
            status_date = stripped_poam_row[POAM_FIELDS.index('STATUS_DATE')]
            vendor_dependency = stripped_poam_row[POAM_FIELDS.index('VENDOR_DEPENDENCY')]
            last_vendor_check_in_date = stripped_poam_row[POAM_FIELDS.index('LAST_VENDOR_CHECK_IN_DATE')]
            vendor_dependent_product_name = stripped_poam_row[POAM_FIELDS.index('VENDOR_DEPENDENT_PRODUCT_NAME')]
            original_risk_rating = stripped_poam_row[POAM_FIELDS.index('ORIGINAL_RISK_RATING')]
            adjusted_risk_rating = stripped_poam_row[POAM_FIELDS.index('ADJUSTED_RISK_RATING')]
            risk_adjustment = stripped_poam_row[POAM_FIELDS.index('RISK_ADJUSTMENT')]
            false_positive = stripped_poam_row[POAM_FIELDS.index('FALSE_POSITIVE')]
            operational_requirement = stripped_poam_row[POAM_FIELDS.index('OPERATIONAL_REQUIREMENT')]
            deviation_rationale = stripped_poam_row[POAM_FIELDS.index('DEVIATION_RATIONALE')]
            supporting_documents = stripped_poam_row[POAM_FIELDS.index('SUPPORTING_DOCUMENTS')]
            comments = stripped_poam_row[POAM_FIELDS.index('COMMENTS')]
            auto_approve = stripped_poam_row[POAM_FIELDS.index('AUTO_APPROVE')]
            poam = POAM(poam_id=poam_id,
                        controls=controls,
                        weakness_name=weakness_name,
                        weakness_description=weakness_description,
                        weakness_detector_source=weakness_detector_source,
                        weakness_source_identifier=weakness_source_identifier,
                        asset_identifier=asset_identifier,
                        point_of_contact=point_of_contact,
                        resources_required=resources_required,
                        overall_remediation_plan=overall_remediation_plan,
                        original_detection_date=original_detection_date,
                        scheduled_completion_date=scheduled_completion_date,
                        planned_milestones=planned_milestones,
                        milestone_changes=milestone_changes,
                        status_date=status_date,
                        vendor_dependency=vendor_dependency,
                        last_vendor_check_in_date=last_vendor_check_in_date,
                        vendor_dependent_product_name=vendor_dependent_product_name,
                        original_risk_rating=original_risk_rating,
                        adjusted_risk_rating=adjusted_risk_rating,
                        risk_adjustment=risk_adjustment,
                        false_positive=false_positive,
                        operational_requirement=operational_requirement,
                        deviation_rationale=deviation_rationale,
                        supporting_documents=supporting_documents,
                        comments=comments,
                        auto_approve=auto_approve
                        )
            poam.create_poam_details()
            poam_entries.append([poam_id, poam])
        result_count = line_number
        # Now we can create the POAM Report object
        # return_poam_report = POAM_Report(in_year=in_year, in_month=in_month, report_name=in_poam_excel_file)
        # return_poam_report.result_count = result_count
        # return_poam_report.poams = poam_entries
        return cls(in_year=in_year, in_month=in_month, report_name=in_poam_excel_file, poams=poam_entries)

    @classmethod
    def poam_from_json_file(cls, in_year, in_month, in_json_poam_file):
        if in_json_poam_file.endswith(".json"):
            logging.info("processing excel file [%s]", in_json_poam_file)
        else:
            logging.error("[%s] is not an JSON file", in_json_poam_file)
            raise NameError(in_json_poam_file)
        try:
            logging.info("Opening POAM JSON file: %s", in_json_poam_file)
            with open(in_json_poam_file, "r") as in_json_poam_fh:
                json_poam_results = json.load(in_json_poam_fh)
        except Exception as e:
            logging.error("Error POAM JSON file: %s", in_json_poam_file)
            raise ValueError("Error POAM JSON file [%s]: [%s]", in_json_poam_file, str(e))
        report_year = in_year
        report_month = in_month
        report_name = str(in_json_poam_file)
        poam_entries = []
        line_number = 0
        for poam_entry in json_poam_results['POAMS']:
            poam = POAM(poam_id=poam_entry['POAM_ID'],
                        controls=poam_entry['CONTROLS'],
                        weakness_name=poam_entry['WEAKNESS_NAME'],
                        weakness_description=poam_entry['WEAKNESS_DESCRIPTION'],
                        weakness_detector_source=poam_entry['WEAKNESS_DETECTOR_SOURCE'],
                        weakness_source_identifier=poam_entry['WEAKNESS_SOURCE_IDENTIFIER'],
                        asset_identifier=poam_entry['ASSET_IDENTIFIER'],
                        point_of_contact=poam_entry['POINT_OF_CONTACT'],
                        resources_required=poam_entry['RESOURCES_REQUIRED'],
                        overall_remediation_plan=poam_entry['OVERALL_REMEDIATION_PLAN'],
                        original_detection_date=poam_entry['ORIGINAL_DETECTION_DATE'],
                        scheduled_completion_date=poam_entry['SCHEDULED_COMPLETION_DATE'],
                        planned_milestones=poam_entry['PLANNED_MILESTONES'],
                        milestone_changes=poam_entry['MILESTONE_CHANGES'],
                        status_date=poam_entry['STATUS_DATE'],
                        vendor_dependency=poam_entry['VENDOR_DEPENDENCY'],
                        last_vendor_check_in_date=poam_entry['LAST_VENDOR_CHECK_IN_DATE'],
                        vendor_dependent_product_name=poam_entry['VENDOR_DEPENDENT_PRODUCT_NAME'],
                        original_risk_rating=poam_entry['ORIGINAL_RISK_RATING'],
                        adjusted_risk_rating=poam_entry['ADJUSTED_RISK_RATING'],
                        risk_adjustment=poam_entry['RISK_ADJUSTMENT'],
                        false_positive=poam_entry['FALSE_POSITIVE'],
                        operational_requirement=poam_entry['OPERATIONAL_REQUIREMENT'],
                        deviation_rationale=poam_entry['DEVIATION_RATIONALE'],
                        supporting_documents=poam_entry['SUPPORTING_DOCUMENTS'],
                        comments=poam_entry['COMMENTS'],
                        auto_approve=poam_entry['AUTO_APPROVE']
                        )
            poam_entries.append(poam)
        result_count = line_number
        # Now we can create the POAM Report object
        # return_poam_report = POAM_Report(in_year=in_year, in_month=in_month, report_name=in_poam_excel_file)
        # return_poam_report.result_count = result_count
        # return_poam_report.poams = poam_entries
        return cls(in_year=report_year, in_month=report_month, report_name=report_name, poams=poam_entries)

    def create_json_poam_results_file(self, out_json_file):
        logging.info("create poam results JSON output from file [%s]", self.report_name)
        current_date_time = datetime.now().strftime('%B %d, %Y %I:%M:%S EDT')
        out_dictionary = dict(REPORT_NAME=self.report_name,
                              SCAN_RESULT_COUNT=self.results_count,
                              SCAN_RESULTS_YEAR=self.report_year,
                              SCAN_RESULTS_MONTH=self.report_month,
                              REPORT_GEN_DATETIME=current_date_time,  # self.in_scan_date, # datetime.now(),
                              POAMS=self.poams)
        # for poam in self.poams:
        #     out_dictionary['POAMS'].append(poam.get_poam_dict())
        logging.info("Generating JSON output file [%s]", out_json_file)
        with open(out_json_file, 'w') as json_out_handle:
            json.dump(out_dictionary, json_out_handle, default=datetime_default)

    def create_poam_excel_report_output(self, excel_report_output_file):
        logging.info("create POAM Excel to file [%s]", excel_report_output_file)
        current_date_time = datetime.datetime.now().strftime('%B %d, %Y %I:%M:%S EDT')
        scan_result_output_wb = openpyxl.load_workbook(filename=POAMReport.FRM_POAM_TEMPLATE_FORM_FULL_PATH,
                                                       keep_vba=True)
        #    scan_result_output_wb = openpyxl.Workbook()
        scan_result_output_ws = scan_result_output_wb[POAMReport.BLANK_POAM_WORKSHEET]
        scan_result_output_ws["D3"] = current_date_time
        poam_result_row = POAMReport.BLANK_POAM_FORM_FIRST_ROW - 1
        # poam_entries = [i[1] for i in self.poams]
        for poam_entry in self.poams:
            poam_result_row += 1
            affected_hosts_str = POAMReport.AFFECTED_HOST_COUNT_STRING.format(host_count=str(len(poam_entry[1].affected_assets)))
            for resource in poam_entry[1].affected_assets:
                affected_hosts_str += POAMReport.AFFECTED_SERVER_STRING.format(resource=resource['NAME'],
                                                                               ip_address=resource['IP_ADDRESS'],
                                                                               port=",".join(resource['PORTS']))
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('POAM_ID') + 1,
                                       value=poam_entry[1].poam_id)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('CONTROLS') + 1,
                                       value=poam_entry[1].controls)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('WEAKNESS_NAME') + 1,
                                       value=poam_entry[1].weakness_name)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('WEAKNESS_DESCRIPTION') + 1,
                                       value=poam_entry[1].weakness_description)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('WEAKNESS_DETECTOR_SOURCE') + 1,
                                       value=poam_entry[1].weakness_detector_source)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('WEAKNESS_SOURCE_IDENTIFIER') + 1,
                                       value=poam_entry[1].weakness_source_identifier)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('ASSET_IDENTIFIER') + 1,
                                       value=affected_hosts_str)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('POINT_OF_CONTACT') + 1,
                                       value=poam_entry[1].point_of_contact)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('RESOURCES_REQUIRED') + 1,
                                       value=poam_entry[1].resources_required)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('OVERALL_REMEDIATION_PLAN') + 1,
                                       value=poam_entry[1].overall_remediation_plan)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('ORIGINAL_DETECTION_DATE') + 1,
                                       value=poam_entry[1].original_detection_date)
            # scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('SCHEDULED_COMPLETION_DATE') + 1,
            #                            value=poam_entry[1].scheduled_completion_date)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('PLANNED_MILESTONES') + 1,
                                       value=poam_entry[1].planned_milestones)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('MILESTONE_CHANGES') + 1,
                                       value=poam_entry[1].milestone_changes)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('STATUS_DATE') + 1,
                                       value=poam_entry[1].status_date)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('VENDOR_DEPENDENCY') + 1,
                                       value=poam_entry[1].vendor_dependency)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('LAST_VENDOR_CHECK_IN_DATE') + 1,
                                       value=poam_entry[1].last_vendor_check_in_date)
            scan_result_output_ws.cell(row=poam_result_row,
                                       column=POAM_FIELDS.index('VENDOR_DEPENDENT_PRODUCT_NAME') + 1,
                                       value=poam_entry[1].vendor_dependent_product_name)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('ORIGINAL_RISK_RATING') + 1,
                                       value=poam_entry[1].original_risk_rating)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('ADJUSTED_RISK_RATING') + 1,
                                       value=poam_entry[1].adjusted_risk_rating)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('RISK_ADJUSTMENT') + 1,
                                       value=poam_entry[1].risk_adjustment)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('FALSE_POSITIVE') + 1,
                                       value=poam_entry[1].false_positive)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('OPERATIONAL_REQUIREMENT') + 1,
                                       value=poam_entry[1].operational_requirement)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('DEVIATION_RATIONALE') + 1,
                                       value=poam_entry[1].deviation_rationale)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('SUPPORTING_DOCUMENTS') + 1,
                                       value=poam_entry[1].supporting_documents)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('COMMENTS') + 1,
                                       value=poam_entry[1].comments)
            scan_result_output_ws.cell(row=poam_result_row, column=POAM_FIELDS.index('AUTO_APPROVE') + 1,
                                       value=poam_entry[1].auto_approve)
        logging.info("Saving POAM Excel to [%s]", excel_report_output_file)
        scan_result_output_wb.save(excel_report_output_file)


if __name__ == "__main__":
    # logging.info("Testing POAM_Report class")
    current_open_poam_file_path = "C:\\Users\\dhartman\\Documents\\FedRAMP\\Continuous Monitoring\\POAM\\FRM\\05 - MAY\\"
    in_poam_file = "FedRAMP-FRM-POAM-May-2020.xlsm"
    open_poam_worksheet_name = "Open POA&M Items"
    closed_poam_worksheet_name = "Closed POA&M Items"
    out_json_filename = "PTC_FRM_CURRENT_POAMS.json"
    current_poam_full_path = os.path.join(current_open_poam_file_path, in_poam_file)
    poam_json_file_out = os.path.join(current_open_poam_file_path, out_json_filename)
