import os
import datetime
import model.CMDBMongoDB as CMDB
import logging
import coloredlogs
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import openpyxl
from mongoengine import *

YEAR_MONTH_DAY = datetime.datetime.now().strftime("%Y%m%d")
ENVIRONMENT = "L5"

FEDRAMP_ROOT_DIR = "C:\\Users\\dhartman\\Documents\\FedRAMP\\CMDB\\"
CMDB_OUTPUT_SPREADSHEET = "L5\\sample-mongodb-output.xlsx"
CMDB_OUTPUT_SPREADSHEET_FULL_PATH = os.path.join(FEDRAMP_ROOT_DIR, CMDB_OUTPUT_SPREADSHEET)

CMDB_WORKSHEET = "CMDB-INVENTORY"
CMDB_NOTES_HISTORY_WORKSHEET = "CMDB_NOTES_HISTORY"
CMDB_IDENTIFIER_MAPPING_WORKSHEET = "NAME_IDENTIFIER_MAPPING"

CMDB_DATA_START_ROW = 2
MONGODB_DATABASE = "CMDB"
CMDB_FIELDS = ['ID',
               'PRIMARY_IP_ADDRESS',
               'NAME',
               'STATUS',
               'ADDITIONAL_IP_ADDRESSES',
               'ENVIRONMENT',
               'FUNCTION',
               'SYSTEM_ADMINISTRATOR_OWNER',
               'APPLICATION_ADMINISTRATOR_OWNER',
               'UNIQUE_ASSET_IDENTIFIER',
               'IPV4_OR_IPV6_ADDRESS',
               'VIRTUAL',
               'PUBLIC',
               'DNS_NAME_OR_URL',
               'NETBIOS_NAME',
               'MAC_ADDRESS',
               'AUTHENTICATED_SCAN',
               'BASELINE_CONFIGURATION_NAME',
               'OS_NAME_AND_VERSION',
               'LOCATION',
               'ASSET_TYPE',
               'HARDWARE_MAKE_MODEL',
               'IN_LATEST_SCAN',
               'SOFTWARE_DATABASE_VENDOR',
               'SOFTWARE_DATABASE_NAME_VERSION',
               'PATCH_LEVEL',
               'COMMENTS',
               'SERIAL_NUMBER_ASSET_TAG_NUMBER',
               'VLAN_NETWORK_ID',
               'CURRENT_FLAG',
               'DELETE_FLAG'
               ]
CMDB_COMMENTS_HISTORY = ['CMDB_RESOURCE_NAME',
                             'COMMENT_NOTE',
                             'TIMESTAMP'
                             ]
CMDB_NAME_IDENTIFIER_MAPPING = ['NAME',
                                'ID',
                                'TYPE',
                                'IDENTIFER'
                                ]
SCAN_POLICY_TARGET = ["SCAN_NAME",
                      "POLICY_NAME",
                      "CREDENTIALS",
                      "TARGETS"]

coloredlogs.install(level=logging.INFO,
                    fmt="%(asctime)s %(hostname)s %(name)s %(filename)s line-%(lineno)d %(levelname)s - %(message)s",
                    datefmt='%H:%M:%S')


def clean_cell(cell):
    if type(cell) == str:
        return cell.strip().upper()
    elif cell is None:
        return ""

    return cell


def create_system_resource_record_from_excel_row(row):
    if row[CMDB_FIELDS.index('STATUS')] == "INACTIVE":
        delete_flag = True
    else:
        delete_flag = False
    logging.info("delete flag [%s] with status [%s] is [%s]", row[CMDB_FIELDS.index('NAME')],
                 row[CMDB_FIELDS.index('STATUS')], str(delete_flag))
    new_cmdb_item = CMDB.SystemResource(name=row[CMDB_FIELDS.index('NAME')],
                                        primary_ip_address=row[CMDB_FIELDS.index('PRIMARY_IP_ADDRESS')],
                                        status=row[CMDB_FIELDS.index('STATUS')],
                                        additional_ip_addresses=row[CMDB_FIELDS.index('ADDITIONAL_IP_ADDRESSES')],
                                        environment=row[CMDB_FIELDS.index('ENVIRONMENT')],
                                        function=row[CMDB_FIELDS.index('FUNCTION')],
                                        system_administrator_owner=row[
                                            CMDB_FIELDS.index('SYSTEM_ADMINISTRATOR_OWNER')],
                                        application_administrator_owner=row[
                                            CMDB_FIELDS.index('APPLICATION_ADMINISTRATOR_OWNER')],
                                        unique_asset_identifier=row[CMDB_FIELDS.index('UNIQUE_ASSET_IDENTIFIER')],
                                        ipv4_or_ipv6_address=row[CMDB_FIELDS.index('IPV4_OR_IPV6_ADDRESS')],
                                        virtual=row[CMDB_FIELDS.index('VIRTUAL')],
                                        public=row[CMDB_FIELDS.index('PUBLIC')],
                                        dns_name_or_url=row[CMDB_FIELDS.index('DNS_NAME_OR_URL')],
                                        netbios_name=row[CMDB_FIELDS.index('NETBIOS_NAME')],
                                        mac_address=row[CMDB_FIELDS.index('MAC_ADDRESS')],
                                        authenticated_scan=row[CMDB_FIELDS.index('AUTHENTICATED_SCAN')],
                                        baseline_configuration_name=row[
                                            CMDB_FIELDS.index('BASELINE_CONFIGURATION_NAME')],
                                        os_name_and_version=row[CMDB_FIELDS.index('OS_NAME_AND_VERSION')],
                                        location=row[CMDB_FIELDS.index('LOCATION')],
                                        asset_type=row[CMDB_FIELDS.index('ASSET_TYPE')],
                                        hardware_make_model=row[CMDB_FIELDS.index('HARDWARE_MAKE_MODEL')],
                                        in_latest_scan=row[CMDB_FIELDS.index('IN_LATEST_SCAN')],
                                        software_database_vendor=row[CMDB_FIELDS.index('SOFTWARE_DATABASE_VENDOR')],
                                        software_database_name_version=row[
                                            CMDB_FIELDS.index('SOFTWARE_DATABASE_NAME_VERSION')],
                                        patch_level=row[CMDB_FIELDS.index('PATCH_LEVEL')],
                                        comments=row[CMDB_FIELDS.index('COMMENTS')],
                                        serial_number_asset_tag_number=row[
                                            CMDB_FIELDS.index('SERIAL_NUMBER_ASSET_TAG_NUMBER')],
                                        vlan_network_id=row[CMDB_FIELDS.index('VLAN_NETWORK_ID')],
                                        delete_flag=delete_flag
                                        )
    return new_cmdb_item


# def compare_sysytem_resource_items(system_resource_current, system_resource_new):
#     change_string = ""
#     item_changed = False
#     # Check individual fields for any changes and prepare a string to add to comment field
#     if stored_cmdb_item.primary_ip_address != new_cmdb_item.primary_ip_address:
#         change_string += "Primary IP Address changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.primary_ip_address,
#                 new=new_cmdb_item.primary_ip_address)
#         stored_cmdb_item.primary_ip_address = new_cmdb_item.primary_ip_address
#         item_changed = True
#     if stored_cmdb_item.additional_ip_addresses != new_cmdb_item.additional_ip_addresses:
#         change_string = "additional_ip_addresses IP Address changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.additional_ip_addresses,
#                 new=new_cmdb_item.additional_ip_addresses)
#         stored_cmdb_item.additional_ip_addresses = new_cmdb_item.additional_ip_addresses
#         item_changed = True
#     if stored_cmdb_item.environment != new_cmdb_item.environment:
#         change_string = "environment changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.environment,
#                 new=new_cmdb_item.environment)
#         stored_cmdb_item.environment = new_cmdb_item.environment
#         item_changed = True
#     if stored_cmdb_item.function != new_cmdb_item.function:
#         change_string = "function changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.function,
#                 new=new_cmdb_item.function)
#         stored_cmdb_item.function = new_cmdb_item.function
#         item_changed = True
#     if stored_cmdb_item.system_administrator_owner != new_cmdb_item.system_administrator_owner:
#         change_string = "system_administrator_owner changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.system_administrator_owner,
#                 new=new_cmdb_item.system_administrator_owner)
#         stored_cmdb_item.system_administrator_owner = new_cmdb_item.system_administrator_owner
#         item_changed = True
#     if stored_cmdb_item.application_administrator_owner != new_cmdb_item.application_administrator_owner:
#         change_string = "application_administrator_owner changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.application_administrator_owner,
#                 new=new_cmdb_item.application_administrator_owner)
#         stored_cmdb_item.application_administrator_owner = new_cmdb_item.application_administrator_owner
#         item_changed = True
#     if stored_cmdb_item.unique_asset_identifier != new_cmdb_item.unique_asset_identifier:
#         change_string = "unique_asset_identifier changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.unique_asset_identifier,
#                 new=new_cmdb_item.unique_asset_identifier)
#         stored_cmdb_item.unique_asset_identifier = new_cmdb_item.unique_asset_identifier
#         item_changed = True
#     if stored_cmdb_item.ipv4_or_ipv6_address != new_cmdb_item.ipv4_or_ipv6_address:
#         change_string = "ipv4_or_ipv6_address changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.ipv4_or_ipv6_address,
#                 new=new_cmdb_item.ipv4_or_ipv6_address)
#         stored_cmdb_item.ipv4_or_ipv6_address = new_cmdb_item.ipv4_or_ipv6_address
#         item_changed = True
#     if stored_cmdb_item.virtual != new_cmdb_item.virtual:
#         change_string = "virtual changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.virtual,
#                 new=new_cmdb_item.virtual)
#         stored_cmdb_item.virtual = new_cmdb_item.virtual
#         item_changed = True
#     if stored_cmdb_item.public != new_cmdb_item.public:
#         change_string += "public changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.public,
#                 new=new_cmdb_item.public)
#         stored_cmdb_item.public = new_cmdb_item.public
#         item_changed = True
#     if stored_cmdb_item.dns_name_or_url != new_cmdb_item.dns_name_or_url:
#         change_string = "dns_name_or_url changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.dns_name_or_url,
#                 new=new_cmdb_item.dns_name_or_url)
#         stored_cmdb_item.dns_name_or_url = new_cmdb_item.dns_name_or_url
#         item_changed = True
#     if stored_cmdb_item.netbios_name != new_cmdb_item.netbios_name:
#         change_string = "netbios_name changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.netbios_name,
#                 new=new_cmdb_item.netbios_name)
#         stored_cmdb_item.netbios_name = new_cmdb_item.netbios_name
#         item_changed = True
#     if stored_cmdb_item.mac_address != new_cmdb_item.mac_address:
#         change_string = "mac_address changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.mac_address,
#                 new=new_cmdb_item.mac_address)
#         stored_cmdb_item.mac_address = new_cmdb_item.mac_address
#         item_changed = True
#     if stored_cmdb_item.authenticated_scan != new_cmdb_item.authenticated_scan:
#         change_string = "authenticated_scan changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.authenticated_scan,
#                 new=new_cmdb_item.authenticated_scan)
#         stored_cmdb_item.authenticated_scan = new_cmdb_item.authenticated_scan
#         item_changed = True
#     if stored_cmdb_item.baseline_configuration_name != new_cmdb_item.baseline_configuration_name:
#         change_string = "baseline_configuration_name changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.baseline_configuration_name,
#                 new=new_cmdb_item.baseline_configuration_name)
#         stored_cmdb_item.baseline_configuration_name = new_cmdb_item.baseline_configuration_name
#         item_changed = True
#     if stored_cmdb_item.os_name_and_version != new_cmdb_item.os_name_and_version:
#         change_string = "os_name_and_version changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.os_name_and_version,
#                 new=new_cmdb_item.os_name_and_version)
#         stored_cmdb_item.os_name_and_version = new_cmdb_item.os_name_and_version
#         item_changed = True
#     if stored_cmdb_item.location != new_cmdb_item.location:
#         change_string = "location changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.location,
#                 new=new_cmdb_item.location)
#         stored_cmdb_item.location = new_cmdb_item.location
#         item_changed = True
#     if stored_cmdb_item.asset_type != new_cmdb_item.asset_type:
#         change_string = "asset_type changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.asset_type,
#                 new=new_cmdb_item.asset_type)
#         stored_cmdb_item.asset_type = new_cmdb_item.asset_type
#         item_changed = True
#     if stored_cmdb_item.hardware_make_model != new_cmdb_item.hardware_make_model:
#         change_string = "hardware_make_model changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.hardware_make_model,
#                 new=new_cmdb_item.hardware_make_model)
#         stored_cmdb_item.hardware_make_model = new_cmdb_item.hardware_make_model
#         item_changed = True
#     if stored_cmdb_item.in_latest_scan != new_cmdb_item.in_latest_scan:
#         change_string = "in_latest_scan changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.in_latest_scan,
#                 new=new_cmdb_item.in_latest_scan)
#         stored_cmdb_item.in_latest_scan = new_cmdb_item.in_latest_scan
#         item_changed = True
#     if stored_cmdb_item.software_database_vendor != new_cmdb_item.software_database_vendor:
#         change_string = "software_database_vendor changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.software_database_vendor,
#                 new=new_cmdb_item.software_database_vendor)
#         stored_cmdb_item.software_database_vendor = new_cmdb_item.software_database_vendor
#         item_changed = True
#     if stored_cmdb_item.software_database_name_version != new_cmdb_item.software_database_name_version:
#         change_string = "software_database_name_version changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.software_database_name_version,
#                 new=new_cmdb_item.software_database_name_version)
#         stored_cmdb_item.software_database_name_version = new_cmdb_item.software_database_name_version
#         item_changed = True
#     if stored_cmdb_item.patch_level != new_cmdb_item.patch_level:
#         change_string = "patch_level changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.patch_level,
#                 new=new_cmdb_item.patch_level)
#         stored_cmdb_item.patch_level = new_cmdb_item.patch_level
#         item_changed = True
#     if stored_cmdb_item.comments != new_cmdb_item.comments:
#         change_string = "comments changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.comments,
#                 new=new_cmdb_item.comments)
#         stored_cmdb_item.comments = new_cmdb_item.comments
#         item_changed = True
#     if stored_cmdb_item.serial_number_asset_tag_number != new_cmdb_item.serial_number_asset_tag_number:
#         change_string = "serial_number_asset_tag_number changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.serial_number_asset_tag_number,
#                 new=new_cmdb_item.serial_number_asset_tag_number)
#         stored_cmdb_item.serial_number_asset_tag_number = new_cmdb_item.serial_number_asset_tag_number
#         item_changed = True
#     if stored_cmdb_item.vlan_network_id != new_cmdb_item.vlan_network_id:
#         change_string = "vlan_network_id changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.vlan_network_id,
#                 new=new_cmdb_item.vlan_network_id)
#         stored_cmdb_item.vlan_network_id = new_cmdb_item.vlan_network_id
#         item_changed = True
#     if stored_cmdb_item.status != new_cmdb_item.status:
#         change_string = "status changed from {old} to {new}\n".format(
#                 old=stored_cmdb_item.status,
#                 new=new_cmdb_item.status)
#         stored_cmdb_item.status = new_cmdb_item.status
#         if stored_cmdb_item.status == "INACTIVE":
#             stored_cmdb_item.delete_flag = True
#         item_changed = True
#
#     if item_changed:
#         logging.info("Record [%s] changed [%s]", new_cmdb_item.name, change_string)
#         new_history_comment = CMDB.Comment(content=change_string)
#         stored_cmdb_item.history_comments.append(new_history_comment)
#
#     return item_changed


def colnum_string(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string


if __name__ == "__main__":
    # Establishing a Connection
    connect('CMDB', host='localhost', port=27017, username="cmdbuser", password="Helen)))1")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=True)
    logging.info("Exporting MongoDB documents to file [%s]", CMDB_OUTPUT_SPREADSHEET_FULL_PATH)
    cmdb_wb = Workbook()
    cmdb_ws = cmdb_wb.create_sheet(CMDB_WORKSHEET)
    cmdb_ws.append(CMDB_FIELDS)
    row_count = 1
    cmdb_item_comments = {}
    cmdb_name_identifer_mapping = {}
    for stored_cmdb_item in CMDB.SystemResource.objects():
        row_count += 1
        row_to_add = [stored_cmdb_item.id,
                      stored_cmdb_item.primary_ip_address,
                      stored_cmdb_item.name,
                      stored_cmdb_item.status,
                      stored_cmdb_item.additional_ip_addresses,
                      stored_cmdb_item.environment,
                      stored_cmdb_item.function,
                      stored_cmdb_item.system_administrator_owner,
                      stored_cmdb_item.application_administrator_owner,
                      stored_cmdb_item.unique_asset_identifier,
                      stored_cmdb_item.ipv4_or_ipv6_address,
                      stored_cmdb_item.virtual,
                      stored_cmdb_item.public,
                      stored_cmdb_item.dns_name_or_url,
                      stored_cmdb_item.netbios_name,
                      stored_cmdb_item.mac_address,
                      stored_cmdb_item.authenticated_scan,
                      stored_cmdb_item.baseline_configuration_name,
                      stored_cmdb_item.os_name_and_version,
                      stored_cmdb_item.location,
                      stored_cmdb_item.asset_type,
                      stored_cmdb_item.hardware_make_model,
                      stored_cmdb_item.in_latest_scan,
                      stored_cmdb_item.software_database_vendor,
                      stored_cmdb_item.software_database_name_version,
                      stored_cmdb_item.patch_level,
                      stored_cmdb_item.comments,
                      stored_cmdb_item.serial_number_asset_tag_number,
                      stored_cmdb_item.vlan_network_id,
                      stored_cmdb_item.current_flag,
                      stored_cmdb_item.delete_flag
                      ]
        cmdb_item_comments[stored_cmdb_item.name] = stored_cmdb_item.history_comments
        cmdb_name_identifer_mapping[stored_cmdb_item.name] = stored_cmdb_item.system_identifiers
        cmdb_ws.append(row_to_add)
    cmdb_inventory_table_ref = "A1:{column}{row}".format(column=colnum_string(len(CMDB_FIELDS)),
                                                         row=row_count)
    cmdb_excel_table = Table(displayName="CMDB", ref=cmdb_inventory_table_ref)
    cmdb_excel_table.tableStyleInfo = style
    cmdb_ws.add_table(cmdb_excel_table)

    # now create a worksheet with the comment/notes history
    cmdb_ws_history = cmdb_wb.create_sheet(CMDB_NOTES_HISTORY_WORKSHEET)
    cmdb_ws_history.append(CMDB_COMMENTS_HISTORY)
    history_row_count = 1
    for id, values in cmdb_item_comments.items():
        for comment_object in values:
            history_row_count += 1
            cmdb_ws_history_row_to_add = [id,
                                          comment_object.content,
                                          str(comment_object.date_added)
                                          ]
            cmdb_ws_history.append(cmdb_ws_history_row_to_add)
    cmdb_ws_history_table_ref = "A1:{column}{row}".format(column=colnum_string(len(CMDB_COMMENTS_HISTORY)),
                                                         row=history_row_count)
    cmdb_comments_excel_table = Table(displayName="CMDB_HISTORY_COMMENTS", ref=cmdb_ws_history_table_ref)
    cmdb_comments_excel_table.tableStyleInfo = style
    cmdb_ws_history.add_table(cmdb_comments_excel_table)

    # now create a worksheet with name/identifier mapping (IPs, DNS, etc)
    cmdb_ws_identifier_mapping = cmdb_wb.create_sheet(CMDB_IDENTIFIER_MAPPING_WORKSHEET)
    cmdb_ws_identifier_mapping.append(CMDB_NAME_IDENTIFIER_MAPPING)
    cmdb_identifier_row_count = 1
    for id, values in cmdb_name_identifer_mapping.items():
        cmdb_identifier_number = 0
        for name_identifier in values:
            cmdb_identifier_row_count += 1
            cmdb_identifier_number += 1
            cmdb_ws_system_identifier_to_add = [id,
                                                str(id) + "-" + str(cmdb_identifier_number),
                                                name_identifier.identifier_type,
                                                name_identifier.identifier
                                                ]
            cmdb_ws_identifier_mapping.append(cmdb_ws_system_identifier_to_add)
    cmdb_ws_identifier_mapping_table_ref = "A1:{column}{row}".format(column=colnum_string(len(CMDB_NAME_IDENTIFIER_MAPPING)),
                                                          row=cmdb_identifier_row_count)
    cmdb_identifier_mapping_excel_table = Table(displayName="CMDB_IDENTIFIER_MAPPING", ref=cmdb_ws_identifier_mapping_table_ref)
    cmdb_identifier_mapping_excel_table.tableStyleInfo = style
    cmdb_ws_identifier_mapping.add_table(cmdb_identifier_mapping_excel_table)

    # table_ref = "A1:AB{row_count}".format(row_count=row_count)
    # cmdb_excel_table = Table(displayName=CMDB_TABLE_NAME, ref=table_ref)
    # cmdb_excel_table.tableStyleInfo = style
    # cmdb_inventory_worksheet.add_table(cmdb_excel_table)
    logging.info("Saving CMDB Inventory export workbook [%s]", CMDB_OUTPUT_SPREADSHEET_FULL_PATH)
    # sheet_to_remove = cmdb_wb.get_sheet_by_name('Sheet1')
    # cmdb_wb.remove_sheet(sheet_to_remove)
    cmdb_wb.save(CMDB_OUTPUT_SPREADSHEET_FULL_PATH)
