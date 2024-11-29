"""This module will load the current OPEN CMDBS into a dictionary object and make available various functions for
retrieving cmdb results.

"""
import os

import coloredlogs
import dateutil
from dateutil.parser import parse
import datetime
import re
import csv
import openpyxl
import pandas
import json
import logging

UNIQUE_ASSET_IDENTIFIER=1
IPV4_OR_IPV6_ADDRESS=2
VIRTUAL=3
PUBLIC=4
DNS_NAME_OR_URL=5
NETBIOS_NAME=6
MAC_ADDRESS=7
AUTHENTICATED_SCAN=8
BASELINE_CONFIGURATION_NAME=9
OS_NAME_AND_VERSION=10
LOCATION=11
ASSET_TYPE=12
HARDWARE_MAKE_MODEL=13
IN_LATEST_SCAN=14
SOFTWARE_DATABASE_VENDOR=15
SOFTWARE_DATABASE_NAME_VERSION=16
PATCH_LEVEL=17
FUNCTION=18
COMMENTS=19
SERIAL_NUMBER_ASSET_TAG_NUMBER=20
VLAN_NETWORK_ID=21
SYSTEM_ADMINISTRATOR_OWNER=22
APPLICATION_ADMINISTRATOR_OWNER=23


ENVIRONMENT = "L5"
YEAR_MONTH_DAY = datetime.datetime.now().strftime("%Y%m%d")

FEDRAMP_CMDBS_ROOT_DIR = "C:\\Users\\dhartman\\Documents\\FedRAMP\\CMDB\\"

# JSON Source File
CMDB_MASTER_JSON_FILE = "{environment}\\PTC-{environment}-CMDB.json".format(environment=ENVIRONMENT)
CMDB_MASTER_JSON_FILE_FULL_PATH = os.path.join(FEDRAMP_CMDBS_ROOT_DIR, CMDB_MASTER_JSON_FILE)

ALL_CMDBS = {}
logging_level = 'INFO'
coloredlogs.install(level=logging_level,
                    fmt="%(asctime)s %(hostname)s %(name)s %(filename)s line-%(lineno)d %(levelname)s - %(message)s",
                    datefmt='%H:%M:%S')
try:
    print("Opening cmdb master JSON file: {}".format(CMDB_MASTER_JSON_FILE_FULL_PATH))
    with open(CMDB_MASTER_JSON_FILE_FULL_PATH, "r") as cmdb_json_fh:
        ALL_CMDBS = json.load(cmdb_json_fh)
except Exception as e:
    print("Error opening cmdb results JSON file: {}".format(CMDB_MASTER_JSON_FILE_FULL_PATH))
    raise ValueError("Error Reading cmdb JSON file [%s]: [%s]", CMDB_MASTER_JSON_FILE_FULL_PATH, str(e))


def get_dict_cmdb_name_ip_listing():
    """Since a resource can have more than one IP, this will return a dictionary that  maps IP and Name

    """
    return_dict = {}
    for key, values in ALL_CMDBS.items():
        name = values['UNIQUE_ASSET_IDENTIFIER']
        for ip in values['IP_ADDRESSES']:
            return_dict_key = "{name}-{ip}".format(name=name, ip=ip)
            return_dict[return_dict_key] = {}
            return_dict[return_dict_key]['NAME'] = name
            return_dict[return_dict_key]['IP_ADDRESS'] = ip
    return return_dict


def get_array_cmdb_poam_server_listing():
    """Returns an array of all cmdb, poam, Server combinations


    """
    return_array = []
    cmdb_server_ip = get_dict_cmdb_name_ip_listing()
    for cmdb_entry in cmdb_server_ip.values():
        return_array.append([cmdb_entry['NAME'], cmdb_entry['IP_ADDRESS']])
    return return_array

NAME_IP_MAPPING = get_array_cmdb_poam_server_listing()
# load_cmdb_file()

# create an array that has the following entries
# use plm-dev-db has an example
# PLM-DEV-DB : PLM-DEV-DB
# PLM-DEV-DB : 192.168.1.101
# PLM-DEV-DB : PLM-DEV-DB.PTCMSCLOUD.COM
# PLM-DEV-DB : Netbios ID
"""
["PLM-DEV-DB": {
    "UNIQUE_ASSET_IDENTIFIER": "PLM-DEV-DB",
    "IPV4_OR_IPV6_ADDRESS": "10.189.213.50",
    "DNS_NAME = "PLM-DEV-DB.MAVERICK.PTCCLOUD.COM"
    },
    <alias>
"""
def generate_alias_mapping():
    alias_mapping = []
    # name_hash = {}
    for key, values in ALL_CMDBS.items():
        if not values['DNS_NAME_OR_URL']:
            dns_entry = values['UNIQUE_ASSET_IDENTIFIER'].upper() + ".MAVERICK.PTCCLOUD.COM"
        else:
            dns_entry = values['DNS_NAME_OR_URL'].upper().strip()
        ip_addresses = values['IP_ADDRESSES']
        resource_name = values['UNIQUE_ASSET_IDENTIFIER']

        alias_mapping.append([key, resource_name])
        for ips in ip_addresses:
            alias_mapping.append([key, ips.strip()])
        alias_mapping.append([key, dns_entry])
    return alias_mapping


NAME_ALIAS_MAPPING = generate_alias_mapping()

CMDB_HEADER_FIELDS = ['CMDB_HEADER_ID',
                      'CMDB_ID',
                      'CONTROLS',
                      'WEAKNESS_NAME',
                      'WEAKNESS_DESCRIPTION',
                      'WEAKNESS_DETECTOR_SOURCE',
                      'WEAKNESS_SOURCE_IDENTIFIER',
                      'POINT_OF_CONTACT',
                      'RESOURCES_REQUIRED',
                      'OVERALL_REMEDIATION_PLAN',
                      'ORIGINAL_DETECTION_DATE',
                      'SCHEDULED_COMPLETION_DATE',
                      'PLANNED_MILESTONES',
                      'MILESTONE_CHANGES',
                      'STATUS_DATE',
                      'VENDOR_DEPENDENCY',
                      'LAST_VENDOR_CHECK_IN_DATE',
                      'VENDOR_DEPENDENT_PRODUCT_NAME',
                      'ORIGINAL_RISK_RATING',
                      'ADJUSTED_RISK_RATING',
                      'RISK_ADJUSTMENT',
                      'FALSE_POSITIVE',
                      'OPERATIONAL_REQUIREMENT',
                      'DEVIATION_RATIONALE',
                      'SUPPORTING_DOCUMENTS',
                      'COMMENTS',
                      'AUTO_APPROVE'
                      ]


def datetime_default(obj):
    if isinstance(obj, (datetime.date, datetime.datetime)):
        return obj.isoformat()


def create_fedramp_inv_and_mapping_workbook():
    import vulnerability_scans.current_vuln_scan as current_vuln_scan
    template_output_file = os.path.join(FEDRAMP_CMDBS_ROOT_DIR, "PTC-System-Inventory-Template.xlsx")
    new_output_file = os.path.join(FEDRAMP_CMDBS_ROOT_DIR, "PTC-{environment}-Inventory-Working-{ymd}.xlsx".format(
            environment=ENVIRONMENT,ymd=YEAR_MONTH_DAY))
    FIRST_DATA_ROW_INV = 2
    logging.info("Writing System Inventory workbook to %s", new_output_file)
    tempate_workbook = openpyxl.load_workbook(filename=template_output_file, data_only=True)
    inventory_ws = tempate_workbook["Inventory"]
    row_number = FIRST_DATA_ROW_INV - 1
    for key, values in ALL_CMDBS.items():
        row_number += 1
        in_latest_scan = "No"
        ip_address_field_out = ""
        for ip in values['IP_ADDRESSES']:
            ip_address_field_out += ip + "\n"
            logging.info("%s is in recent scan => %s", ip, current_vuln_scan.ip_in_scan_results(ip))
            if current_vuln_scan.ip_in_scan_results(ip):
                in_latest_scan = "Yes"
        inventory_ws.cell(row=row_number, column=UNIQUE_ASSET_IDENTIFIER,
                          value=values['UNIQUE_ASSET_IDENTIFIER'])
        inventory_ws.cell(row=row_number, column=IPV4_OR_IPV6_ADDRESS, value=ip_address_field_out)
        inventory_ws.cell(row=row_number, column=VIRTUAL, value=values['VIRTUAL'])
        inventory_ws.cell(row=row_number, column=PUBLIC, value=values['PUBLIC'])
        inventory_ws.cell(row=row_number, column=DNS_NAME_OR_URL, value=values['DNS_NAME_OR_URL'])
        inventory_ws.cell(row=row_number, column=NETBIOS_NAME, value=values['NETBIOS_NAME'])
        inventory_ws.cell(row=row_number, column=MAC_ADDRESS, value=values['MAC_ADDRESS'])
        inventory_ws.cell(row=row_number, column=AUTHENTICATED_SCAN, value=values['AUTHENTICATED_SCAN'])
        inventory_ws.cell(row=row_number, column=BASELINE_CONFIGURATION_NAME,
                          value=values['BASELINE_CONFIGURATION_NAME'])
        inventory_ws.cell(row=row_number, column=OS_NAME_AND_VERSION, value=values['OS_NAME_AND_VERSION'])
        # inventory_ws.cell(row=row_number, column=OS_NAME_AND_VERSION, value=values['OS_NAME_VERSION'])
        inventory_ws.cell(row=row_number, column=LOCATION, value=values['LOCATION'])
        inventory_ws.cell(row=row_number, column=ASSET_TYPE, value=values['ASSET_TYPE'])
        inventory_ws.cell(row=row_number, column=HARDWARE_MAKE_MODEL, value=values['HARDWARE_MAKE_MODEL'])
        inventory_ws.cell(row=row_number, column=IN_LATEST_SCAN, value=in_latest_scan)
        inventory_ws.cell(row=row_number, column=SOFTWARE_DATABASE_VENDOR,
                          value=values['SOFTWARE_DATABASE_VENDOR'])
        inventory_ws.cell(row=row_number, column=SOFTWARE_DATABASE_NAME_VERSION,
                          value=values['SOFTWARE_DATABASE_NAME_VERSION'])
                          # value=values['SOFTWARE_DB_NAME_VERSION'])
        inventory_ws.cell(row=row_number, column=PATCH_LEVEL, value=values['PATCH_LEVEL'])
        inventory_ws.cell(row=row_number, column=FUNCTION, value=values['FUNCTION'])
        inventory_ws.cell(row=row_number, column=COMMENTS, value=values['COMMENTS'])
        inventory_ws.cell(row=row_number, column=SERIAL_NUMBER_ASSET_TAG_NUMBER,
                          value=values['SERIAL_NUMBER_ASSET_TAG_NUMBER'])
                          # value=values['SERIAL_NUM_ASSET_TAG'])
        inventory_ws.cell(row=row_number, column=VLAN_NETWORK_ID, value=values['VLAN_NETWORK_ID'])
        inventory_ws.cell(row=row_number, column=SYSTEM_ADMINISTRATOR_OWNER,
                          # value=values['SYSTEM_ADMIN_OWNER'])
                          value=values['SYSTEM_ADMINISTRATOR_OWNER'])
        inventory_ws.cell(row=row_number, column=APPLICATION_ADMINISTRATOR_OWNER,
                          # value=values['APP_ADMIN_OWNER'])
                          value=values['APPLICATION_ADMINISTRATOR_OWNER'])

    tempate_workbook.save(filename=new_output_file)
    logging.info("System Inventory saved to %s", new_output_file)


def print_name_ip_listing():
    """Prints tab delimieted output of name and IP

    """
    cmdb_server_ip = get_dict_cmdb_name_ip_listing()
    for cmdb_entry in cmdb_server_ip.values():
        print("{name}\t{ip}".format(name=cmdb_entry['NAME'],
                                    ip=cmdb_entry['IP_ADDRESS']))


def print_name_ip_string():
    """Prints tab delimieted output of name and IP

    """
    cmdb_server_ip = get_dict_cmdb_name_ip_listing()
    for cmdb_entry in cmdb_server_ip.values():
        print("{name} ({ip})".format(name=cmdb_entry['NAME'],
                                    ip=cmdb_entry['IP_ADDRESS']))


def get_name_from_ip(ip_in):
    return_name = ""
    for key, values in ALL_CMDBS.items():
        for entry in NAME_IP_MAPPING:
            name = entry[0]
            ip = entry[1]
            if ip == ip_in:
                return_name = name
    return return_name


def get_array_cmdb_poam_server_listing():
    """Returns an array of all cmdb, poam, Server combinations


    """
    return_array = []
    cmdb_server_ip = get_dict_cmdb_name_ip_listing()
    for cmdb_entry in cmdb_server_ip.values():
        return_array.append([cmdb_entry['NAME'], cmdb_entry['IP_ADDRESS']])
    return return_array


def clipboard_copy_cmdb_poam_server_listing():
    """Returns an array of all cmdb, poam, Server combinations


    """
    return_array = []
    cmdb_server_ip = get_dict_cmdb_name_ip_listing()
    for cmdb_entry in cmdb_server_ip.values():
        return_array.append([cmdb_entry['NAME'], cmdb_entry['IP_ADDRESS']])
    df = pandas.DataFrame(return_array)
    df.to_clipboard(index=False, header=False)


def print_cmdb_to_csv(csv_output_file):
    # Skip file body if there is nothing to write to file
    cmdb_out_array = [[]]
    for values in ALL_CMDBS.values():
        for ip in values['IP_ADDRESSES']:
            cmdb_out_array.append([values['UNIQUE_ASSET_IDENTIFIER'],
                                   values['IN_LATEST_SCAN'],
                                   values['DNS_NAME_OR_URL'],
                                   ip])
    my_df = pandas.DataFrame(cmdb_out_array)
    my_df.to_csv(csv_output_file, index=False, header=False)
    # try:
    #     with open(csv_output_file, 'w') as csv_out_handle:
    #         # Write header row
    #         # csv_out_handle.write(COMMA_SEPARATOR.join(header_columns) + '\n')
    #         for cmdb_key, cmdb_item in cmdb_result_dict.items():
    #             # Get all of the dictionary attributes
    #             print("cmdb Key: " + str(cmdb_key) +
    #                   " item: " + str(cmdb_item))
    #             fedramp_inv_ipv4_or_ipv6_address = cmdb_item['IPV4_OR_IPV6_ADDRESS']
    #
    #             fedramp_inv_application_administrator_owner = cmdb_result_dict[
    #                 'APPLICATION_ADMINISTRATOR_OWNER']
    #             # Generate a row for each IP/Name
    #             for ip_name in fedramp_inv_ipv4_or_ipv6_address:
    #                 out_row = [cmdb_item['UNIQUE_ASSET_IDENTIFIER'],
    #                            ip_name,
    #                            cmdb_item['VIRTUAL'],
    #                            cmdb_item['PUBLIC'],
    #                            cmdb_item['DNS_NAME_OR_URL'],
    #                            cmdb_item['NETBIOS_NAME'],
    #                            cmdb_item['MAC_ADDRESS'],
    #                            cmdb_item['AUTHENTICATED_SCAN'],
    #                            cmdb_item['BASELINE_CONFIGURATION_NAME'],
    #                            cmdb_item['OS_NAME_AND_VERSION'],
    #                            cmdb_item['LOCATION'],
    #                            cmdb_item['ASSET_TYPE'],
    #                            cmdb_item['HARDWARE_MAKE_MODEL'],
    #                            cmdb_item['IN_LATEST_SCAN'],
    #                            cmdb_item['SOFTWARE_DATABASE_VENDOR'],
    #                            cmdb_item['SOFTWARE_DATABASE_NAME_VERSION'],
    #                            cmdb_item['PATCH_LEVEL'],
    #                            cmdb_item['FUNCTION'],
    #                            cmdb_item['COMMENTS'],
    #                            cmdb_item['SERIAL_NUMBER_ASSET_TAG'],
    #                            cmdb_item['VLAN_NETWORK_ID'],
    #                            cmdb_item['SYSTEM_ADMINISTRATOR_OWNER'],
    #                            cmdb_item['APPLICATION_ADMINISTRATOR_OWNER']
    #                            ]
    #                 csv_out_handle.write(COMMA_SEPARATOR.join(out_row) + '\n')
    # except Exception as e:
    #     print("Exception: " + str(e) +
    #           "\n  Error occurred while writing to TSV file " + str(csv_output_file))


def get_name_from_alias(in_alias):
    return_name = ""
    for row in NAME_ALIAS_MAPPING:
        if row[1].upper() == in_alias.upper():
            return_name = row[0]
    if return_name == "":
        logging.warn("No alias mapping record found for [%s]", in_alias)
    return return_name

if __name__ == "__main__":
    create_fedramp_inv_and_mapping_workbook()